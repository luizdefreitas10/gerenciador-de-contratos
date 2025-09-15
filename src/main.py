import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Caminho da planilha dentro da pasta src
BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # pega pasta do main.py
PLANILHA_PATH = os.path.join(BASE_DIR, "Planilha de Contratos 2025 - CÓPIA.xlsx")

# Carregar planilha com openpyxl para poder editar (cores)
wb = load_workbook(PLANILHA_PATH)
ws = wb["2024"]  # aba dos contratos
gestores_ws = wb["Gestores"]  # aba com os emails

# Criar um dataframe da aba 2024 (linha 5 como cabeçalho)
df_contratos = pd.read_excel(PLANILHA_PATH, sheet_name="2024", header=4)

# Função para pegar email do gestor pelo nome
def get_email_gestor(nome_gestor: str) -> str:
    df_gestores = pd.read_excel(PLANILHA_PATH, sheet_name="Gestores", header=0)
    linha = df_gestores[df_gestores["nome"].str.lower() == nome_gestor.lower()]
    if not linha.empty:
        return linha["email"].values[0]
    return None

# Configuração do servidor de email
SMTP_SERVER = "smtp.seuservidor.com"
SMTP_PORT = 587
EMAIL_USER = "seuemail@dominio.com"
EMAIL_PASS = "suasenha"

def enviar_email(destinatario, assunto, mensagem):
    msg = MIMEMultipart()
    msg["From"] = EMAIL_USER
    msg["To"] = destinatario
    msg["Subject"] = assunto

    msg.attach(MIMEText(mensagem, "plain"))

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASS)
        server.send_message(msg)

# Preencher cor vermelha para contratos próximos do vencimento
fill_red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

hoje = datetime.today()

for idx, row in df_contratos.iterrows():
    fim_vigencia = row["FIM DA VIGÊNCIA DO CONTRATO [18]"]
    fiscal = row["NOMES DO FISCAL CONTRATO [20]"]

    if pd.notnull(fim_vigencia):
        dias_restantes = (fim_vigencia - hoje).days

        if dias_restantes <= 90:
            # Pintar a linha no Excel
            excel_row = idx + 6  # pq header está na linha 5 → dados começam na 6
            for cell in ws[excel_row]:
                cell.fill = fill_red

            # Buscar email do gestor
            email = get_email_gestor(fiscal)
            if email:
                mensagem = f"""
                Prezado {fiscal},

                O contrato "{row['NOME DO CONTRATO']}" está a {dias_restantes} dias do fim da vigência ({fim_vigencia.date()}).

                Favor tomar as devidas providências.
                """
                enviar_email(email, "Alerta: Contrato próximo ao vencimento", mensagem)

# Salvar planilha com formatação aplicada
wb.save(PLANILHA_PATH)
print("Processo concluído com sucesso 🚀")
